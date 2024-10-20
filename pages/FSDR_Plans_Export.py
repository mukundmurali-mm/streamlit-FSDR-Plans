import streamlit as st
import oci
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font
from io import BytesIO
from commonLib import *

st.set_page_config(
    page_title="FSDR Plans Export",
    page_icon="☁️"  # Cloud emoji to represent OCI
)

def main():
    st.title("FSDR Plans Export")
    

    if 'oci_profile' not in st.session_state:
        st.error("Please select an OCI profile on the main page.")
        return

    # Input fields
    ocid = st.text_input("DR Plan OCID")
    sheet_name = st.text_input("Sheet Name")
    
    # Excel file name input
    use_default_filename = st.checkbox("Use default file name", value=True)
    if use_default_filename:
        file_name = "dr_plan_export.xlsx"
    else:
        file_name = st.text_input("Enter Excel file name", value="dr_plan_export.xlsx")
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'

    if st.button("Export Plan"):
        try:
            # Load region map
            region_file = os.path.join(os.path.dirname(__file__), "region_file.json")
            region_map = load_region_map(region_file)
            region = get_region_from_ocid(ocid, region_map)

            # Set up OCI config and signer
            config = oci.config.from_file(profile_name=st.session_state['oci_profile'])
            config['region'] = region
            signer = oci.signer.Signer.from_config(config)

            # Get DR Plan
            disaster_recovery_client = oci.disaster_recovery.DisasterRecoveryClient(
                config=config, retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY, signer=signer)
            get_dr_plan_response = disaster_recovery_client.get_dr_plan(dr_plan_id=ocid)
            plan_groups = get_dr_plan_response.data.plan_groups

            # Extract the order of plan groups
            original_order = [pg.id for pg in plan_groups]

            # Manually convert DrPlanGroup objects to dictionaries
            plan_dicts = []
            for pg in plan_groups:
                steps = []
                for step in pg.steps:
                    step_dict = {
                        'display_name': step.display_name,
                        'error_mode': step.error_mode,
                        'id': step.id,
                        'is_enabled': step.is_enabled,
                        'timeout': step.timeout,
                        'type': step.type,
                    }
                    if hasattr(step, 'user_defined_step') and step.user_defined_step:
                        user_defined_step = {
                            'step_type': step.user_defined_step.step_type,
                            'run_as_user': getattr(step.user_defined_step, 'run_as_user', None),
                            'run_on_instance_id': getattr(step.user_defined_step, 'run_on_instance_id', None),
                            'function_id': getattr(step.user_defined_step, 'function_id', None),
                            'function_region': getattr(step.user_defined_step, 'function_region', None),
                            'request_body': getattr(step.user_defined_step, 'request_body', None),
                            'object_storage_script_location': {
                                'bucket': getattr(step.user_defined_step.object_storage_script_location, 'bucket', None),
                                'namespace': getattr(step.user_defined_step.object_storage_script_location, 'namespace', None),
                                'object': getattr(step.user_defined_step.object_storage_script_location, 'object', None)
                            } if getattr(step.user_defined_step, 'object_storage_script_location', None) else None,
                            'run_on_instance_region': getattr(step.user_defined_step, 'run_on_instance_region', None),
                            'script_command': getattr(step.user_defined_step, 'script_command', None)
                        }
                        step_dict['user_defined_step'] = user_defined_step
                    steps.append(step_dict)
                plan_dicts.append({
                    'display_name': pg.display_name,
                    'id': pg.id,
                    'type': pg.type,
                    'steps': steps,
                    'is_pause_enabled': pg.is_pause_enabled
                })

            # Convert the parsed plan data to a DataFrame
            df = pd.json_normalize(plan_dicts)

            # Split the data into two parts based on the "type" value
            built_in_df = df[(df['type'] == 'BUILT_IN')]
            other_df = df[df['type'] != 'BUILT_IN']
            pause_in_df = df[df['type'] == 'USER_DEFINED_PAUSE']

            # Function to normalize and reformat data
            def normalize_and_reformat(df):
                dict_list_orient = df.to_dict('records')
                normalized_data = pd.json_normalize(dict_list_orient, "steps", ['display_name', 'id', 'type'], record_prefix='steps.')
                columns_order = [
                    'display_name', 'id', 'steps.display_name', 'steps.error_mode', 'steps.id', 'steps.is_enabled',
                    'steps.timeout', 'steps.type', 'type'
                ]
                normalized_data = normalized_data.reindex(columns=columns_order, fill_value=None)
                return normalized_data

            def pause(df):
                dict_list_orient = df.to_dict('records')
                normalized_data = pd.json_normalize(dict_list_orient)
                columns_order = [
                    'display_name', 'id', 'is_pause_enabled', 'type'
                ]
                normalized_data = normalized_data.reindex(columns=columns_order, fill_value=None)
                return normalized_data

            def normalize_other_data(df):
                dict_list_orient = df.to_dict('records')
                normalized_data = pd.json_normalize(dict_list_orient, "steps", ['display_name', 'id', 'type'], record_prefix='steps.')
                columns_order = [
                    'display_name', 'id', 'steps.display_name', 'steps.error_mode', 'steps.id', 'steps.is_enabled',
                    'steps.timeout', 'steps.type', 'steps.user_defined_step.step_type',
                    'steps.user_defined_step.run_as_user', 'steps.user_defined_step.run_on_instance_id',
                    'steps.user_defined_step.function_id', 'steps.user_defined_step.function_region', 'steps.user_defined_step.request_body',
                    'steps.user_defined_step.object_storage_script_location.bucket', 'steps.user_defined_step.object_storage_script_location.namespace', 'steps.user_defined_step.object_storage_script_location.object',
                    'steps.user_defined_step.run_on_instance_region', 'steps.user_defined_step.script_command', 'type'
                ]
                normalized_data = normalized_data.reindex(columns=columns_order, fill_value=None)
                return normalized_data

            # Normalize and reformat both subsets of data
            built_in_data = normalize_and_reformat(built_in_df)
            other_data = normalize_other_data(other_df)
            pause_data = pause(pause_in_df)

            # Append both subsets of data into one DataFrame
            combined_data = pd.concat([other_data, built_in_data, pause_data], ignore_index=True)

            # Sort the combined data based on the original order
            combined_data['sort_order'] = pd.Categorical(combined_data['id'], categories=original_order, ordered=True)
            combined_data.sort_values('sort_order', inplace=True)
            combined_data.drop(columns=['sort_order'], inplace=True)

            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                combined_data.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                # Apply styling
                def merge_and_center(ws, col):
                    max_row = ws.max_row
                    for row in range(2, max_row + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        start_row = row
                        while row <= max_row and ws.cell(row=row, column=col).value == cell_value:
                            row += 1
                        end_row = row - 1
                        if start_row != end_row:
                            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                            merged_cell = ws.cell(row=start_row, column=col)
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                columns_to_merge = ['A', 'B']

                for col in columns_to_merge:
                    col_index = column_index_from_string(col)
                    merge_and_center(worksheet, col_index)

                # Define fill colors
                fill_blue = PatternFill(start_color="346EC9", end_color="346EC9", fill_type="solid")
                fill_purple = PatternFill(start_color="858491", end_color="858491", fill_type="solid")
                font_white = Font(color="FFFFFF", bold=True)

                header_cells = worksheet[1]
                for cell in header_cells:
                    if cell.column_letter in ['A', 'B', 'T']:
                        cell.fill = fill_blue
                        cell.font = font_white
                    else:
                        cell.fill = fill_purple
                        cell.font = font_white

                # Auto-adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

            # Offer the Excel file for download
            st.download_button(
                label="Download Excel file",
                data=output.getvalue(),
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
