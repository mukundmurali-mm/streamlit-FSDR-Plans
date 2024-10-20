import streamlit as st
import oci
import openpyxl
import os
from commonLib import *
from io import BytesIO

st.set_page_config(
    page_title="FSDR Plans Update",
    page_icon="☁️"  # Cloud emoji to represent OCI
)

def get_merged_cell_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_cell_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_cell_range:
            merged_cell = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
            return merged_cell.value
    return cell.value

def new_plan(row, plan_groups_dict):
    plan_group_display_name = str(row[0])
    id = None
    step_display_name = str(row[2])
    step_error_mode = row[3]
    s_id = None
    step_is_enabled = row[5]
    timeout = row[6]
    step_type = row[8]
    run_as_user = row[9]
    run_on_instance_id = row[10]
    function_id = row[11]
    function_region = row[12]
    request_body = row[13]
    bucket = row[14]
    namespace = row[15]
    bucket_object = row[16]
    instance_region = row[17]
    script_command = row[18]

    if step_type in ["RUN_LOCAL_SCRIPT", "RUN_OBJECTSTORE_SCRIPT", "INVOKE_FUNCTION"]:
        type = 'USER_DEFINED'
    else:
        raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

    valid_step_types = [
        'RUN_OBJECTSTORE_SCRIPT_PRECHECK',
        'RUN_LOCAL_SCRIPT_PRECHECK',
        'INVOKE_FUNCTION_PRECHECK',
        'RUN_OBJECTSTORE_SCRIPT',
        'RUN_LOCAL_SCRIPT',
        'INVOKE_FUNCTION'
    ]

    if step_type not in valid_step_types:
        raise ValueError(f"Invalid step_type: {step_type}. Must be one of {valid_step_types}")

    if plan_group_display_name in plan_groups_dict:
        plan_group_details = plan_groups_dict[plan_group_display_name]
    else:
        plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
            display_name=plan_group_display_name,
            id=id,
            type=type,
            steps=[]
        )
        plan_groups_dict[plan_group_display_name] = plan_group_details

    if step_type == "RUN_LOCAL_SCRIPT":
        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled,
            user_defined_step=oci.disaster_recovery.models.UpdateRunLocalScriptUserDefinedStepDetails(
                step_type=step_type,
                run_on_instance_id=run_on_instance_id,
                run_as_user=run_as_user,
                script_command=script_command
            )
        )
    elif step_type == "RUN_OBJECTSTORE_SCRIPT":
        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled,
            user_defined_step=oci.disaster_recovery.models.UpdateRunObjectStoreScriptUserDefinedStepDetails(
                step_type=step_type,
                run_on_instance_id=run_on_instance_id,
                object_storage_script_location=oci.disaster_recovery.models.UpdateObjectStorageScriptLocationDetails(
                    bucket=bucket,
                    namespace=namespace,
                    object=bucket_object
                )
            )
        )
    elif step_type == "INVOKE_FUNCTION":
        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled,
            user_defined_step=oci.disaster_recovery.models.UpdateInvokeFunctionUserDefinedStepDetails(
                step_type=step_type,
                function_id=function_id,
                request_body=request_body
            )
        )
    else:
        raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

    if step_details not in plan_group_details.steps:
        plan_group_details.steps.append(step_details)

    return plan_groups_dict, plan_group_details

def existing_plan(row, plan_groups_dict):
    plan_group_display_name = str(row[0])
    id = str(row[1])
    step_display_name = str(row[2])
    step_error_mode = row[3]
    s_id = str(row[4])
    step_is_enabled = row[5]
    timeout = row[6]
    step_type = row[8]
    run_as_user = row[9]
    run_on_instance_id = row[10]
    function_id = row[11]
    function_region = row[12]
    request_body = row[13]
    bucket = row[14]
    namespace = row[15]
    bucket_object = row[16]
    instance_region = row[17]
    script_command = row[18]

    if step_type in ["RUN_LOCAL_SCRIPT", "RUN_OBJECTSTORE_SCRIPT", "INVOKE_FUNCTION"]:
        type = 'USER_DEFINED'
    else:
        raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

    valid_step_types = [
        'RUN_OBJECTSTORE_SCRIPT_PRECHECK',
        'RUN_LOCAL_SCRIPT_PRECHECK',
        'INVOKE_FUNCTION_PRECHECK',
        'RUN_OBJECTSTORE_SCRIPT',
        'RUN_LOCAL_SCRIPT',
        'INVOKE_FUNCTION'
    ]

    if step_type not in valid_step_types:
        raise ValueError(f"Invalid step_type: {step_type}. Must be one of {valid_step_types}")

    if id in plan_groups_dict:
        plan_group_details = plan_groups_dict[id]
    else:
        plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
            display_name=plan_group_display_name,
            id=id,
            type=type,
            steps=[]
        )
        plan_groups_dict[id] = plan_group_details

    if step_type == "RUN_LOCAL_SCRIPT":
        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled,
            user_defined_step=oci.disaster_recovery.models.UpdateRunLocalScriptUserDefinedStepDetails(
                step_type=step_type,
                run_on_instance_id=run_on_instance_id,
                run_as_user=run_as_user,
                script_command=script_command
            )
        )
    elif step_type == "RUN_OBJECTSTORE_SCRIPT":
        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled,
            user_defined_step=oci.disaster_recovery.models.UpdateRunObjectStoreScriptUserDefinedStepDetails(
                step_type=step_type,
                run_on_instance_id=run_on_instance_id,
                object_storage_script_location=oci.disaster_recovery.models.UpdateObjectStorageScriptLocationDetails(
                    bucket=bucket,
                    namespace=namespace,
                    object=bucket_object
                )
            )
        )
    elif step_type == "INVOKE_FUNCTION":
        step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
            display_name=step_display_name,
            error_mode=step_error_mode,
            id=s_id,
            timeout=timeout,
            is_enabled=step_is_enabled,
            user_defined_step=oci.disaster_recovery.models.UpdateInvokeFunctionUserDefinedStepDetails(
                step_type=step_type,
                function_id=function_id,
                request_body=request_body
            )
        )
    else:
        raise ValueError(f"Invalid step_type: {step_type}. Must be one of RUN_LOCAL_SCRIPT, RUN_OBJECTSTORE_SCRIPT, INVOKE_FUNCTION")

    if step_details not in plan_group_details.steps:
        plan_group_details.steps.append(step_details)

    return plan_groups_dict, plan_group_details

def pause_plan(row, plan_groups_dict):
    plan_group_display_name = str(row[0])
    id = str(row[1])
    type = 'USER_DEFINED_PAUSE'

    if id in plan_groups_dict:
        plan_group_details = plan_groups_dict[id]
    else:
        plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
            display_name=plan_group_display_name,
            id=id,
            type=type,
            is_pause_enabled=True
        )
        plan_groups_dict[id] = plan_group_details

    return plan_groups_dict, plan_group_details

def builtin_function(row, plan_groups_dict):
    plan_group_display_name = str(row[0])
    id = str(row[1])
    step_display_name = str(row[2])
    step_error_mode = row[3]
    s_id = row[4]
    step_is_enabled = row[5]
    timeout = row[6]
    type = row[19]

    valid_builtin_types = ['BUILT_IN', 'BUILT_IN_PRECHECK', 'USER_DEFINED', 'USER_DEFINED_PAUSE']
    if type not in valid_builtin_types:
        raise ValueError(f"Invalid value for `type`: {type}. Must be one of {valid_builtin_types}")

    if id in plan_groups_dict:
        plan_group_details = plan_groups_dict[id]
    else:
        plan_group_details = oci.disaster_recovery.models.UpdateDrPlanGroupDetails(
            display_name=plan_group_display_name,
            id=id,
            type=type,
            steps=[]
        )
        plan_groups_dict[id] = plan_group_details

    step_details = oci.disaster_recovery.models.UpdateDrPlanStepDetails(
        display_name=step_display_name,
        error_mode=step_error_mode,
        id=s_id,
        timeout=timeout,
        is_enabled=step_is_enabled
    )

    if step_details not in plan_group_details.steps:
        plan_group_details.steps.append(step_details)

    return plan_groups_dict, plan_group_details

def main():
    st.title("FSDR Plans Update")

    if 'oci_profile' not in st.session_state:
        st.error("Please select an OCI profile on the main page.")
        return

    # Input fields
    ocid = st.text_input("DR Plan OCID")
    sheet_name = st.text_input("Sheet Name")
    excel_file = st.file_uploader("Upload Excel File", type="xlsx")

    if st.button("Update Plan"):
        try:
            # Load region map
            region_file = os.path.join(os.path.dirname(__file__), "region_file.json")
            region_map = load_region_map(region_file)
            region = get_region_from_ocid(ocid, region_map)

            # Set up OCI config and signer
            config = oci.config.from_file(profile_name=st.session_state['oci_profile'])
            config['region'] = region
            signer = oci.signer.Signer.from_config(config)

            # Initialize Disaster Recovery client
            disaster_recovery_client = oci.disaster_recovery.DisasterRecoveryClient(config=config, signer=signer)

            # Load Excel file
            workbook = openpyxl.load_workbook(BytesIO(excel_file.read()))
            sheet = workbook[sheet_name]

            # Process Excel data
            plan_groups_dict = {}
            ordered_plan_groups = []

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                row_values = [get_merged_cell_value(sheet, row[0].row, col) for col in range(1, sheet.max_column + 1)]
                id_value = str(row_values[1])
                type_value = str(row_values[19])

                row_values = [None if val in ["None", None] else val for val in row_values]

                if type_value == "USER_DEFINED":
                    if id_value == "None":
                        plan_groups_dict, plan_group_details = new_plan(row_values, plan_groups_dict)
                    else:
                        plan_groups_dict, plan_group_details = existing_plan(row_values, plan_groups_dict)
                elif type_value == "USER_DEFINED_PAUSE":
                    plan_groups_dict, plan_group_details = pause_plan(row_values, plan_groups_dict)
                else:
                    plan_groups_dict, plan_group_details = builtin_function(row_values, plan_groups_dict)

                ordered_plan_groups.append(plan_group_details)

            final_plan_groups = list(plan_groups_dict.values())

            # Update DR Plan
            update_dr_plan_details = oci.disaster_recovery.models.UpdateDrPlanDetails(plan_groups=final_plan_groups)
            update_dr_plan_response = disaster_recovery_client.update_dr_plan(
                update_dr_plan_details=update_dr_plan_details,
                dr_plan_id=ocid
            )
            st.success(f"Update to DR Plan {ocid} is successful")

        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
